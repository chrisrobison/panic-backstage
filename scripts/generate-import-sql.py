#!/usr/bin/env python3
"""
generate-import-sql.py

Reads MabEvents.xlsx (project root) and regenerates
backstage/database/mabevents-import.sql.

Usage:
    python3 backstage/scripts/generate-import-sql.py

Requires openpyxl:
    pip3 install openpyxl

Tracker sheet column map (row 2 is the header; data starts row 3+):
   1  Priority Level / external_id        → events.external_id
   2  Referral/Provenance                 → events.referral_source
   3  Organizer/promoter                  → events.promoter_name
   4  descrip/genre                       → events.title + event_type guess
   5  Date                                → events.date
   6  Potential Revenue (rent + bar)      → events.potential_revenue
   7  Time: Load in                       → event_schedule_items('load_in')
   8  Time: doors open                    → events.doors_time
   9  Time door clos                      → events.end_time
  10  Time: lock up                       → event_schedule_items('curfew')
  11  Est. crd size                       → events.capacity
  12  Venue (Upstairs/Downstairs/Both)    → events.room
  13  Status                              → events.status (via STATUS_MAP)
  14  Event Type (Public/Private)         → events.public_visibility + event_type
  15  Ticket Sys.                         → events.ticket_system
  16  Contract Link                       → events.contract_url
  17  Walk Through Happened?              → events.walkthrough_done
  18  Ticket Link                         → events.ticket_url
  19  Night of Settlement Document        → events.settlement_doc_url
  20  Bar / freeform notes                → events.description_internal
  21-23 Security / Sound / Cleaning       → ignored (~always empty)

Staff Contact sheet:
   1 Department  → maps to staff_members.default_role
   2 Fname       → staff_members.name (combined)
   3 Lname       → staff_members.name (combined)
   4 Pronoun     → staff_members.pronoun
   5 Staffing Status → ignored (sheet column is empty in practice)
   6 Phone       → staff_members.phone
   7 Email       → staff_members.email + users.email
   8 Position    → staff_members.position
   9 Staffing Notes → staff_members.notes
"""

import openpyxl
import json
import os
import re
import sys
from datetime import datetime, date, time
from pathlib import Path

SCRIPT_DIR   = Path(__file__).resolve().parent
BACKSTAGE    = SCRIPT_DIR.parent
PROJECT_ROOT = BACKSTAGE.parent
XLSX_PATH    = PROJECT_ROOT / 'MabEvents.xlsx'
OUTPUT_SQL   = BACKSTAGE / 'database' / 'mabevents-import.sql'

# Hidden "App ID" column in the sheet (Tracker!Z) carries the immutable app
# event id — the stable link key used to match a sheet row to its event,
# immune to in-app title/date edits. 0-based index 25.
APP_ID_COL_INDEX = 25

# Baseline of last-synced sheet values per event, written by dump-sheet-shadow.php.
# Used for field-level change detection: a field is only pushed into the DB when
# its sheet value changed since the last sync (sheet wins); otherwise the app's
# value is preserved. Empty/missing on first run -> existing linked events are
# "seeded" (shadow recorded, app data untouched).
SHADOW_JSON_PATH = os.environ.get(
    'SHEET_SHADOW_JSON', str(BACKSTAGE / 'storage' / 'tmp' / 'sheet-shadow.json')
)
try:
    with open(SHADOW_JSON_PATH, 'r') as _fh:
        SHADOW_BASELINE = json.load(_fh) or {}
except (FileNotFoundError, json.JSONDecodeError):
    SHADOW_BASELINE = {}

# Pending sheet->app links: events that were CREATED from a new sheet row on a
# previous sync but whose App ID hasn't been confirmed back into the sheet yet
# (written by dump-sheet-import-links.php). For such a row — still blank in the
# sheet's App ID column — we REUSE the already-created event instead of inserting
# again, so a failed/retried write-back can't spawn a duplicate. Indexed both by
# sheet row number (the precise anchor) and by (title, date) as a fallback for
# when rows shifted position between syncs.
LINKS_JSON_PATH = os.environ.get(
    'SHEET_IMPORT_LINKS_JSON', str(BACKSTAGE / 'storage' / 'tmp' / 'sheet-import-links.json')
)
PENDING_BY_ROW: dict[int, dict] = {}
PENDING_BY_TITLE_DATE: dict[tuple, dict] = {}
try:
    with open(LINKS_JSON_PATH, 'r') as _fh:
        for _rec in (json.load(_fh) or []):
            PENDING_BY_ROW[int(_rec['sheet_row'])] = _rec
            PENDING_BY_TITLE_DATE[(str(_rec.get('title') or ''), _rec.get('date'))] = _rec
except (FileNotFoundError, json.JSONDecodeError, KeyError, TypeError, ValueError):
    PENDING_BY_ROW = {}
    PENDING_BY_TITLE_DATE = {}

# ── helpers ──────────────────────────────────────────────────────────────────

def esc(v):
    if v is None:
        return 'NULL'
    v = str(v).replace('\\', '\\\\').replace("'", "\\'").strip()
    return f"'{v}'"

def fmt_time(v):
    if isinstance(v, time):
        return f"'{v.strftime('%H:%M:%S')}'"
    if isinstance(v, datetime):
        return f"'{v.strftime('%H:%M:%S')}'"
    return 'NULL'

def slugify(s: str) -> str:
    s = re.sub(r'[^a-z0-9]+', '-', s.lower().strip())
    return re.sub(r'-+', '-', s).strip('-')

STATUS_MAP = {
    'booked': 'confirmed',
    'in negotiations': 'hold',
    'cancelled': 'canceled',
    'canceled': 'canceled',
    'prospect': 'proposed',
    'paid deposit': 'confirmed',
    'paid in full': 'confirmed',
    'archived': 'completed',
    'hold': 'hold',
}

# Sheet's "Department" → staff_members.default_role enum
DEPT_TO_ROLE = {
    'bar':       'bartender',
    'barback':   'barback',
    'security':  'security',
    'door':      'door',
    'sound':     'sound',
    'light':     'lighting',
    'lighting':  'lighting',
    'stage':     'stagehand',
    'stagehand': 'stagehand',
    'cleaning':  'cleaner',
    'cleaner':   'cleaner',
    'manager':   'manager',
    'runner':    'runner',
}

def map_status(raw_status, ext_id_raw=None) -> str:
    # Status comes from the Status column. (Column A used to be "Priority Level"
    # and could say "hold"; it now holds the EVT-N code, so that heuristic is
    # retired — the Status column is authoritative.)
    if raw_status is None:
        return 'proposed'
    return STATUS_MAP.get(str(raw_status).strip().lower(), 'proposed')

def guess_event_type(genre, event_type_raw) -> str:
    if event_type_raw == 'Private':
        return 'private_event'
    g = (genre or '').lower()
    if 'karaoke' in g:
        return 'karaoke'
    if any(k in g for k in ['comedy', 'clown']):
        return 'comedy'
    if any(k in g for k in [' dj ', 'dj-', 'dj set', 'dj event']):
        return 'dj_night'
    if 'open mic' in g:
        return 'open_mic'
    if any(k in g for k in ['bachata', 'swing lesson', 'swing dancing', 'dance class']):
        return 'special_event'
    return 'live_music'

def clean_ext_id(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if re.match(r'^EVT-\d+', s) else None

def parse_revenue(v):
    """
    Best-effort numeric extraction from the sheet's "Potential Revenue" column.
    Handles plain numbers (2000, 1500.0) and lossy expressions like '500+700'
    or '$1,200' by summing all numbers found in the string. Returns float or None.
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    nums = re.findall(r'\d+(?:\.\d+)?', s.replace(',', ''))
    if not nums:
        return None
    try:
        return sum(float(n) for n in nums)
    except Exception:
        return None

def parse_ticket_system(v):
    """Normalize ticket-system text and clip to schema length (40)."""
    if v is None:
        return None
    s = str(v).strip()
    return s[:40] if s else None

def parse_url_or_text(v, max_len=500):
    """
    The sheet's link columns aren't always URLs — sometimes they're notes
    like 'Verbal contract' or 'text group'. Keep whatever text is there
    (trimmed and clipped) so the data lands somewhere visible; the UI just
    renders this as text and only auto-links when it looks like a URL.
    """
    if v is None:
        return None
    s = str(v).strip()
    return s[:max_len] if s else None

def parse_bool_yes(v):
    """'Yes' / 'TRUE' / '1' → 1, everything else → 0."""
    if v is None:
        return 0
    s = str(v).strip().lower()
    return 1 if s in ('yes', 'y', 'true', '1', 'done', 'complete') else 0

# ── Load workbook ─────────────────────────────────────────────────────────────

if not XLSX_PATH.exists():
    print(f"ERROR: {XLSX_PATH} not found.", file=sys.stderr)
    sys.exit(1)

wb = openpyxl.load_workbook(str(XLSX_PATH))

# ── Parse Tracker ─────────────────────────────────────────────────────────────

ws = wb['Tracker']
used_slugs: set[str] = set()

def unique_slug(base: str) -> str:
    slug = slugify(base)[:90]
    if slug not in used_slugs:
        used_slugs.add(slug)
        return slug
    i = 2
    while f"{slug}-{i}" in used_slugs:
        i += 1
    final = f"{slug}-{i}"
    used_slugs.add(final)
    return final

events_rows = []
skipped = 0

# The header row is row 2 in the live sheet; first data row is row 3 (sometimes
# row 4 — row 3 may be a stray comment). iter_rows(min_row=2) keeps the legacy
# behaviour: rows that fail the date/genre guards are skipped, so a stray
# header passes through harmlessly.
# enumerate(start=2): iter_rows(min_row=2) yields sheet row 2 (the header) first,
# so `sheet_row` is the true 1-based row number — the same numbering app-id-sync.php
# uses to address cells. It anchors the App ID write-back to the exact source row.
for sheet_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    cols = list(row[:20]) + [None] * max(0, 20 - len(row))
    (raw_ext_id, referral, organizer, genre, raw_date, potential_rev,
        t_load_in, t_doors, t_close, t_lockup, crowd_size, venue_room,
        status_raw, event_type_raw, ticket_sys, contract_link, walkthrough_raw,
        ticket_link, settlement_doc, bar_notes) = cols

    # Hidden App ID column (immutable link key); absent for brand-new sheet rows.
    app_id_cell = row[APP_ID_COL_INDEX] if len(row) > APP_ID_COL_INDEX else None
    app_id = None
    if app_id_cell is not None and str(app_id_cell).strip() != '':
        try:
            app_id = int(float(str(app_id_cell).strip()))
        except (ValueError, TypeError):
            app_id = None

    if not isinstance(raw_date, (datetime, date)):
        skipped += 1
        continue
    if not genre and not organizer:
        skipped += 1
        continue

    title    = (str(genre).strip() if genre else str(organizer).strip())[:200]
    date_str = raw_date.strftime('%Y-%m-%d')
    slug     = unique_slug(title + '-' + date_str)

    ext_id = clean_ext_id(raw_ext_id)
    status = map_status(status_raw, raw_ext_id)
    etype  = guess_event_type(genre, event_type_raw)
    pub    = 1 if event_type_raw == 'Public' else 0

    room = None
    if venue_room:
        r = str(venue_room).strip().lower()
        if 'upstairs' in r:
            room = 'upstairs'
        elif 'downstairs' in r:
            room = 'downstairs'
        elif 'both' in r:
            room = 'both'

    # description_internal is now ONLY for stuff that doesn't have its own
    # column — col 20 freeform "Bar" notes. Everything else (revenue, ticket
    # system, contract, settlement doc) goes into its own structured field
    # so we stop polluting the notes blob.
    desc_int = str(bar_notes).strip() if bar_notes and str(bar_notes).strip() else None

    cap = None
    if crowd_size is not None:
        try:
            cap = int(float(str(crowd_size).replace(',', '').split('/')[0].strip()))
        except Exception:
            pass

    events_rows.append({
        'app_id': app_id,
        'sheet_row': sheet_row,
        'title': title, 'slug': slug, 'etype': etype, 'status': status,
        'desc_int': desc_int, 'date': date_str,
        'doors': fmt_time(t_doors), 'end_t': fmt_time(t_close),
        'cap': str(cap) if cap else 'NULL', 'pub': pub,
        'ext_id': ext_id,
        'ref': str(referral).strip() if referral else None,
        'prom': str(organizer).strip() if organizer else None,
        'room': room,
        'load_in': fmt_time(t_load_in),
        'lockup': fmt_time(t_lockup),
        # New structured fields
        'potential_revenue': parse_revenue(potential_rev),
        'ticket_url':        parse_url_or_text(ticket_link),
        'ticket_system':     parse_ticket_system(ticket_sys),
        'contract_url':      parse_url_or_text(contract_link),
        'walkthrough_done':  parse_bool_yes(walkthrough_raw),
        'settlement_doc_url': parse_url_or_text(settlement_doc),
    })

# ── Parse Staff Contact ───────────────────────────────────────────────────────

ws2 = wb['Staff Contact']
staff_rows = []

for row in ws2.iter_rows(min_row=2, values_only=True):
    cols = list(row[:9]) + [None] * max(0, 9 - len(row))
    dept, fname, lname, pronoun, staffing_status, phone, email, position, notes_col = cols
    if not fname:
        continue
    full = (str(fname).strip() + (' ' + str(lname).strip() if lname else '')).strip()
    if not full:
        continue
    real_email = bool(email)
    if not email:
        email = re.sub(r'[^a-z0-9]', '.', full.lower()).strip('.') + '@staff.mabuhay.local'

    dept_norm = (str(dept).strip().lower() if dept else '')
    role = DEPT_TO_ROLE.get(dept_norm, 'other')

    staff_rows.append({
        'name':       full,
        'email':      str(email).strip(),
        'real_email': real_email,
        'dept':       str(dept).strip() if dept else None,
        'phone':      str(phone).strip() if phone else None,
        'pronoun':    str(pronoun).strip() if pronoun else None,
        'position':   str(position).strip() if position else None,
        'notes':      str(notes_col).strip() if notes_col else None,
        'role':       role,
    })

# ── Build SQL ─────────────────────────────────────────────────────────────────

out = [
    "-- =============================================================",
    "-- MabEvents.xlsx import (idempotent UPSERT)",
    "-- Run AFTER: schema.sql, migrations 001 → 011",
    f"-- Events: {len(events_rows)}  |  Staff: {len(staff_rows)}",
    "--",
    "-- Idempotency keys:",
    "--   users         : email (UNIQUE)        — name refreshed; role/password preserved",
    "--   staff_members : email (when present)  — refresh phone/pronoun/position/notes",
    "--   events        : slug  (UNIQUE)        — most fields refreshed; status preserved",
    "--                                          (local workflow state wins)",
    "--   event_schedule_items : delete+reinsert per event for ('load_in','curfew')",
    "-- =============================================================",
    "",
    "START TRANSACTION;",
    "",
    "-- Resolve venue and default owner",
    "SET @venue_id = (SELECT id FROM venues WHERE slug = 'mabuhay-gardens' LIMIT 1);",
    "-- Owner: legacy seed admin if present, else the lowest-id venue_admin.",
    "SET @owner_id = COALESCE(",
    "  (SELECT id FROM users WHERE email = 'admin@mabuhay.local' LIMIT 1),",
    "  (SELECT id FROM users WHERE role = 'venue_admin' ORDER BY id LIMIT 1)",
    ");",
    "",
    "-- ── Staff users ────────────────────────────────────────────────",
]

for s in staff_rows:
    out.append(
        f"INSERT INTO users (name, email, password_hash, role) VALUES "
        f"({esc(s['name'])}, {esc(s['email'])}, NULL, 'staff') "
        f"ON DUPLICATE KEY UPDATE name = VALUES(name);"
    )

out.extend(["", "-- ── Staff roster (staff_members) ─────────────────────────────"])

# Staff roster UPSERT. Idempotency: email is UNIQUE in users but NOT in
# staff_members, so we use a "INSERT ... ON DUPLICATE KEY" pattern via
# a synthetic key — we resolve the existing row by (name, email) match
# and UPDATE if present, INSERT otherwise. To keep this SQL declarative
# we lean on the unique-index-free path: try INSERT, then UPDATE WHERE
# email = ? (which is unique-enough for the sheet's data shape).
for s in staff_rows:
    out.append(
        f"INSERT INTO staff_members (name, email, phone, pronoun, default_role, position, notes, active, user_id) "
        f"SELECT {esc(s['name'])}, {esc(s['email'])}, {esc(s['phone'])}, "
        f"{esc(s['pronoun'])}, '{s['role']}', {esc(s['position'])}, {esc(s['notes'])}, 1, "
        f"(SELECT id FROM users WHERE email = {esc(s['email'])} LIMIT 1) "
        f"FROM DUAL WHERE NOT EXISTS (SELECT 1 FROM staff_members WHERE email = {esc(s['email'])});"
    )
    out.append(
        f"UPDATE staff_members SET "
        f"name = {esc(s['name'])}, "
        f"phone = COALESCE({esc(s['phone'])}, phone), "
        f"pronoun = COALESCE({esc(s['pronoun'])}, pronoun), "
        f"position = COALESCE({esc(s['position'])}, position), "
        f"notes = COALESCE({esc(s['notes'])}, notes), "
        f"user_id = COALESCE((SELECT id FROM users WHERE email = {esc(s['email'])} LIMIT 1), user_id) "
        f"WHERE email = {esc(s['email'])};"
    )

out.extend(["", "-- ── Events ─────────────────────────────────────────────────────"])

# Columns refreshed on the slug-keyed INSERT path, used ONLY for brand-new sheet
# rows that don't carry an App ID yet. Existing, App-ID-linked rows go through
# the field-level merge below instead (which decides per field whether the sheet
# changed). deposit_amount is omitted (no sheet column).
EVENT_UPDATE_COLS = [
    'venue_id', 'title', 'event_type', 'status', 'description_internal',
    'date', 'doors_time', 'end_time', 'capacity', 'public_visibility',
    'owner_user_id', 'referral_source', 'promoter_name', 'room',
    'potential_revenue', 'ticket_url', 'ticket_system', 'contract_url',
    'walkthrough_done', 'settlement_doc_url',
]
event_update_clause = ',\n    '.join(
    [f"{c} = VALUES({c})" for c in EVENT_UPDATE_COLS]
)

def num_or_null(v):
    return 'NULL' if v is None else f"{v}"

def frag_clean(frag):
    """Canonical compare-string for a SQL value fragment ('NULL'->'', strip quotes)."""
    if frag is None:
        return ''
    f = str(frag)
    if f == 'NULL':
        return ''
    return f.strip().strip("'")

def field_specs(ev):
    """[(db_col, compare_str, sql_expr)] for the sheet-owned event-table fields."""
    def s(x):
        return '' if x is None else str(x).strip()
    rev = num_or_null(ev['potential_revenue'])
    return [
        ('title',               s(ev['title']),               esc(ev['title'])),
        ('event_type',          s(ev['etype']),               f"'{ev['etype']}'"),
        ('status',              s(ev['status']),              f"'{ev['status']}'"),
        ('description_internal', s(ev['desc_int']),           esc(ev['desc_int'])),
        ('date',                s(ev['date']),                f"'{ev['date']}'"),
        ('doors_time',          frag_clean(ev['doors']),      ev['doors']),
        ('end_time',            frag_clean(ev['end_t']),      ev['end_t']),
        ('capacity',            frag_clean(ev['cap']),        ev['cap']),
        ('public_visibility',   str(ev['pub']),               str(ev['pub'])),
        # external_id (the EVT-N code) is app-owned and pushed app->sheet; never
        # imported, so a sheet edit to column A can't change an event's code.
        ('referral_source',     s(ev['ref']),                 esc(ev['ref'])),
        ('promoter_name',       s(ev['prom']),                esc(ev['prom'])),
        ('room',                s(ev['room']),                esc(ev['room'])),
        ('potential_revenue',   frag_clean(rev),              rev),
        ('ticket_url',          s(ev['ticket_url']),          esc(ev['ticket_url'])),
        ('ticket_system',       s(ev['ticket_system']),       esc(ev['ticket_system'])),
        ('contract_url',        s(ev['contract_url']),        esc(ev['contract_url'])),
        ('walkthrough_done',    str(ev['walkthrough_done']),  str(ev['walkthrough_done'])),
        ('settlement_doc_url',  s(ev['settlement_doc_url']),  esc(ev['settlement_doc_url'])),
    ]

# By default, unlinked sheet rows (no App ID) are reported but NOT inserted.
# Auto-inserting on slug-drift is what historically spawned duplicate events,
# and during the migration new events should be created in the app. Set
# SHEET_INSERT_NEW=1 to restore the old auto-insert behaviour.
INSERT_NEW = os.environ.get('SHEET_INSERT_NEW', '0') == '1'

def link_upsert(eid_expr, ev) -> str:
    """SQL recording (event_id, sheet_row) in sheet_import_links so the App ID
    write-back can target this row precisely. `linked` is left untouched on
    conflict so a confirmed link is never re-opened."""
    date_sql = f"'{ev['date']}'" if ev['date'] else 'NULL'
    return (
        f"INSERT INTO sheet_import_links (event_id, sheet_row, title_snap, date_snap, linked) "
        f"VALUES ({eid_expr}, {ev['sheet_row']}, {esc(ev['title'])}, {date_sql}, 0) "
        f"ON DUPLICATE KEY UPDATE sheet_row = VALUES(sheet_row), "
        f"title_snap = VALUES(title_snap), date_snap = VALUES(date_snap);"
    )

stat = {'inserted': 0, 'updated': 0, 'unchanged': 0, 'seeded': 0,
        'unlinked': 0, 'relinked': 0}
unlinked_rows = []

# Slug reassignments for already-linked events, deferred to a two-phase
# rename after the main loop (see below) instead of being written inline —
# see the note at the emission site for why.
pending_slug_updates: list[tuple[int, str]] = []

for ev in events_rows:
    specs = field_specs(ev)
    shadow_now = {col: comp for (col, comp, _sql) in specs}
    shadow_now['load_in'] = frag_clean(ev['load_in'])
    shadow_now['curfew']  = frag_clean(ev['lockup'])
    app_id = ev['app_id']
    sched_changed = False

    if app_id is not None:
        base = SHADOW_BASELINE.get(str(app_id))
        if base is None:
            # SEED: linked event with no baseline yet. Record the shadow but do
            # NOT touch the row — the app/DB value is authoritative.
            out.append(f"SET @eid = {app_id};")
            stat['seeded'] += 1
        else:
            changed = [(col, sql) for (col, comp, sql) in specs
                       if col in base and comp != base[col]]
            sched_changed = any(
                k in base and shadow_now[k] != base[k] for k in ('load_in', 'curfew')
            )
            if changed:
                set_parts = [f"{col} = {sql}" for col, sql in changed]
                names = {col for col, _ in changed}
                if 'title' in names or 'date' in names:
                    # Don't set slug inline here: this event's new slug can be
                    # another row's *current* (not-yet-updated) slug — e.g. a
                    # block of sheet edits that shuffles several events' titles
                    # at once. Setting it immediately makes success depend on
                    # this statement happening to come after whichever other
                    # statement vacates that slug, which is fragile ordering
                    # this generator doesn't otherwise guarantee. Deferred to
                    # the two-phase rename after the main loop instead.
                    pending_slug_updates.append((app_id, ev['slug']))
                out.append(f"UPDATE events SET {', '.join(set_parts)} WHERE id = {app_id};")
                stat['updated'] += 1
            else:
                stat['unchanged'] += 1
            out.append(f"SET @eid = {app_id};")
    else:
        # No App ID in the sheet for this row. Two cases:
        pending = (PENDING_BY_ROW.get(ev['sheet_row'])
                   or PENDING_BY_TITLE_DATE.get((ev['title'], ev['date'])))
        if pending is not None:
            # We ALREADY created an event from this row on a prior sync; its App
            # ID just hasn't been confirmed back into the sheet yet. Reuse that
            # event — never insert a second one — and keep the link tracked so
            # the write-back step retries. This is what makes the create rule
            # safe to retry across a failed write-back (no duplicate-on-drift).
            reuse_id = int(pending['event_id'])
            out.append(f"SET @eid = {reuse_id};")
            out.append(link_upsert(str(reuse_id), ev))
            stat['relinked'] += 1
        elif not INSERT_NEW:
            # Unlinked sheet row: report it, don't insert.
            unlinked_rows.append(f"{ev['title']} ({ev['date']})")
            out.append(f"-- unlinked sheet row (not imported): {ev['title']} ({ev['date']})")
            stat['unlinked'] += 1
            continue
        else:
            # Brand-new sheet row (no App ID): slug-keyed upsert to create the
            # event, then record (event_id, sheet_row) in sheet_import_links so
            # app-id-sync.php link-imports writes its id back into THIS exact row.
            out.append(
                f"INSERT INTO events (venue_id, title, slug, event_type, status, "
                f"description_internal, date, doors_time, end_time, capacity, "
                f"public_visibility, owner_user_id, external_id, referral_source, "
                f"promoter_name, room, potential_revenue, ticket_url, ticket_system, "
                f"contract_url, walkthrough_done, settlement_doc_url) VALUES ("
                f"@venue_id, {esc(ev['title'])}, {esc(ev['slug'])}, "
                f"'{ev['etype']}', '{ev['status']}', {esc(ev['desc_int'])}, "
                f"'{ev['date']}', {ev['doors']}, {ev['end_t']}, {ev['cap']}, "
                f"{ev['pub']}, @owner_id, {esc(ev['ext_id'])}, "
                f"{esc(ev['ref'])}, {esc(ev['prom'])}, {esc(ev['room'])}, "
                f"{num_or_null(ev['potential_revenue'])}, "
                f"{esc(ev['ticket_url'])}, {esc(ev['ticket_system'])}, "
                f"{esc(ev['contract_url'])}, {ev['walkthrough_done']}, "
                f"{esc(ev['settlement_doc_url'])}"
                f")\n  ON DUPLICATE KEY UPDATE\n    id = LAST_INSERT_ID(id),\n    "
                f"{event_update_clause};"
            )
            out.append("SET @eid = LAST_INSERT_ID();")
            out.append(link_upsert('@eid', ev))
            sched_changed = True
            stat['inserted'] += 1

    # Schedule items (load_in / curfew): only rewrite when they changed in the
    # sheet (or it's a new row), so app-edited schedules aren't reverted.
    if sched_changed:
        out.append(
            "DELETE FROM event_schedule_items "
            "WHERE event_id = @eid AND item_type IN ('load_in','curfew');"
        )
        if ev['load_in'] != 'NULL':
            out.append(
                f"INSERT INTO event_schedule_items (event_id, title, item_type, start_time) "
                f"VALUES (@eid, 'Load-in', 'load_in', {ev['load_in']});"
            )
        if ev['lockup'] != 'NULL':
            out.append(
                f"INSERT INTO event_schedule_items (event_id, title, item_type, start_time) "
                f"VALUES (@eid, 'Lock-up / Curfew', 'curfew', {ev['lockup']});"
            )

    # Record/refresh this event's shadow (keyed by the resolved @eid) so the
    # next sync can tell which sheet fields changed.
    shadow_json = json.dumps(shadow_now, ensure_ascii=True, sort_keys=True)
    out.append(
        f"INSERT INTO event_sheet_shadow (event_id, raw_json, synced_at) "
        f"VALUES (@eid, {esc(shadow_json)}, NOW()) "
        f"ON DUPLICATE KEY UPDATE raw_json = VALUES(raw_json), synced_at = VALUES(synced_at);"
    )

if pending_slug_updates:
    out.extend([
        "",
        "-- ── Slug renames (two-phase) ──────────────────────────────────",
        "-- `slug` is UNIQUE, so a rename can collide with another row's",
        "-- CURRENT slug value mid-transaction — either a legitimate cascade",
        "-- (several events' titles shuffled in one sheet edit) or a stale",
        "-- row this sync doesn't otherwise touch. Phase 1 vacates every",
        "-- renaming event to a placeholder derived from its own id (always",
        "-- unique, so it can never collide with anything). Phase 2 then",
        "-- assigns the real slugs, order-independent since every event in",
        "-- this batch has already given up its old slug. A genuine conflict",
        "-- (the target slug still held by an untouched row) correctly fails",
        "-- phase 2 rather than being silently papered over.",
    ])
    for app_id, _new_slug in pending_slug_updates:
        out.append(f"UPDATE events SET slug = CONCAT('__pending_slug__', id) WHERE id = {app_id};")
    for app_id, new_slug in pending_slug_updates:
        out.append(f"UPDATE events SET slug = {esc(new_slug)} WHERE id = {app_id};")

out.extend(["", "COMMIT;"])

sql = '\n'.join(out)
OUTPUT_SQL.write_text(sql)

print(f"Written: {OUTPUT_SQL}")
print(f"Events:  {len(events_rows)}  (skipped {skipped} rows without a parseable date)")
print(f"Merge:   {stat['updated']} updated, {stat['unchanged']} unchanged, "
      f"{stat['inserted']} new, {stat['relinked']} relinked, {stat['seeded']} seeded, "
      f"{stat['unlinked']} unlinked (baseline: {len(SHADOW_BASELINE)} events)")
if unlinked_rows:
    print(f"Unlinked sheet rows (reported, not imported — set SHEET_INSERT_NEW=1 to import):")
    for r in unlinked_rows:
        print(f"   • {r}")
print(f"Staff:   {len(staff_rows)}")
